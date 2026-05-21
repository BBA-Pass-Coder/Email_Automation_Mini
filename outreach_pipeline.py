"""
Root Labs / Be Bodywise — Hair Serum Creator Outreach Pipeline

End-to-end automation:
  Input:  CSV with creator handles + email + bio + niche
  Output: CSV with subject + body + confidence + reasoning per creator

Flow per creator:
  1. Pull recent videos via Apify TikTok scraper
  2. Score and pick the best video (spoken, hair-relevant, recent, high views)
  3. Get transcript via Fusion MCP (analyze_tiktok_video)
  4. Draft email via Claude Sonnet 4 with the locked template
  5. Write all results to output CSV

Setup:
  pip install apify-client anthropic
  export APIFY_TOKEN=...
  export ANTHROPIC_API_KEY=...

Run:
  python outreach_pipeline.py --input creators_input.csv --output creators_output.csv
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional

from apify_client import ApifyClient
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# =========================
# Config
# =========================

APIFY_TOKEN = os.environ.get("APIFY_TOKEN")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")

if not APIFY_TOKEN:
    sys.exit("Set APIFY_TOKEN environment variable.")
if not ANTHROPIC_API_KEY:
    sys.exit("Set ANTHROPIC_API_KEY environment variable.")

APIFY_ACTOR = "clockworks/tiktok-scraper"
CLAUDE_MODEL = "claude-sonnet-4-5"
FUSION_MCP_URL = "https://fusion-production-ff7f.up.railway.app/mcp"

VIDEOS_PER_CREATOR = 10  # how many recent videos to fetch per profile
MAX_VIDEO_AGE_DAYS = 90

HAIR_KEYWORDS = [
    "hair", "wig", "serum", "scalp", "rosemary", "redensyl", "anagain",
    "edges", "growth", "thickness", "thinning", "bald", "follicle", "lace",
    "frontal", "install", "wash day", "curl", "curly", "natural hair",
    "4c", "4a", "4b", "3c", "3b", "3a", "blowout", "silk press", "extension",
    "weave", "braid", "twist", "loc", "haircut", "hairstyle", "stylist",
    "cosmetologist", "trichologist", "shedding", "breakage", "moisture"
]

# =========================
# Step 1: Fetch videos via Apify
# =========================

def fetch_videos_for_handles(handles: list[str]) -> dict[str, list[dict]]:
    """
    Calls clockworks/tiktok-scraper with a batch of profiles.
    Returns {handle: [video_dict, ...]} for handles that returned data.
    """
    client = ApifyClient(APIFY_TOKEN)

    # Strip @ from handles, scraper expects bare usernames
    clean_handles = [h.lstrip("@").strip() for h in handles]

    run_input = {
        "profiles": clean_handles,
        "resultsPerPage": VIDEOS_PER_CREATOR,
        "shouldDownloadCovers": False,
        "shouldDownloadVideos": False,
        "shouldDownloadSubtitles": False,
    }

    print(f"[Apify] Scraping {len(clean_handles)} profiles...")
    run = client.actor(APIFY_ACTOR).call(run_input=run_input)

    items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
    print(f"[Apify] Got {len(items)} videos total")

    by_handle = {}
    for item in items:
        # Scraper returns nested authorMeta dict with the username
        author = (item.get("authorMeta") or {}).get("name") or ""
        author = author.lower()
        if not author:
            continue
        by_handle.setdefault(author, []).append(item)

    return by_handle


# =========================
# Step 2: Score and pick best video
# =========================

def score_video(video: dict) -> float:
    score = 0.0

    # Caption hair-relevance
    caption = (video.get("text") or "").lower()
    keyword_hits = sum(1 for kw in HAIR_KEYWORDS if kw in caption)
    score += keyword_hits * 8

    # Duration: prefer 15-120s (long enough to speak, short enough to be a real post)
    meta = video.get("videoMeta") or {}
    duration = meta.get("duration", 0) or 0
    if 15 <= duration <= 120:
        score += 25
    elif 10 <= duration < 15 or 120 < duration <= 180:
        score += 10

    # Recency
    created_ts = video.get("createTimeISO") or video.get("createTime")
    if created_ts:
        try:
            if isinstance(created_ts, str):
                created = datetime.fromisoformat(created_ts.replace("Z", "+00:00"))
            else:
                created = datetime.fromtimestamp(created_ts, tz=timezone.utc)
            days_old = (datetime.now(timezone.utc) - created).days
            if days_old > MAX_VIDEO_AGE_DAYS:
                return -1  # disqualify stale
            if days_old < 30:
                score += 15
            elif days_old < 60:
                score += 8
        except Exception:
            pass

    # Views (log-ish, capped)
    views = video.get("playCount") or 0
    if views > 0:
        import math
        score += min(math.log10(views) * 5, 20)

    return score


def select_best_video(videos: list[dict]) -> Optional[dict]:
    if not videos:
        return None
    scored = [(v, score_video(v)) for v in videos]
    scored = [s for s in scored if s[1] >= 0]
    if not scored:
        return None
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored[0][0]


# =========================
# Step 3: Transcript via Fusion MCP
# =========================

def get_transcript(video_url: str, anthropic: Anthropic) -> dict:
    """
    Calls Fusion's analyze_tiktok_video tool through Claude with MCP.
    Returns {'transcript': str, 'visual_summary': str} or {'error': str}.
    """
    try:
        message = anthropic.beta.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=4000,
            mcp_servers=[{
                "type": "url",
                "url": FUSION_MCP_URL,
                "name": "fusion",
            }],
            messages=[{
                "role": "user",
                "content": (
                    f"Use the Fusion analyze_tiktok_video tool to fetch the transcript "
                    f"and visual analysis for this video: {video_url}\n\n"
                    f"Then return a JSON object with two fields:\n"
                    f'  "transcript": the full spoken transcript as a single string\n'
                    f'  "visual_summary": a 2-3 sentence summary of what is visually happening\n\n'
                    f"Return ONLY the JSON, no preamble."
                )
            }],
            extra_headers={"anthropic-beta": "mcp-client-2025-04-04"},
        )

        # Pull text content from response
        text_blocks = [b.text for b in message.content if hasattr(b, "text")]
        full_text = "\n".join(text_blocks).strip()

        # Strip code fences if present
        full_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", full_text, flags=re.MULTILINE).strip()

        parsed = json.loads(full_text)
        return {
            "transcript": parsed.get("transcript", ""),
            "visual_summary": parsed.get("visual_summary", ""),
        }
    except Exception as e:
        return {"error": str(e)}


# =========================
# Step 4: Draft the outreach email
# =========================

EMAIL_PROMPT = """You are drafting a single outreach email from Mayank, CEO of Root Labs and Be Bodywise, to a TikTok creator. The product is a hair growth serum.

USE THIS EXACT TEMPLATE. Only customize the [X] and [Y] portions of the first paragraph. Everything else stays verbatim.

---
Hi {first_name},

I am Mayank, CEO of Root Labs and Be Bodywise. Came across your [X] on TikTok and liked [Y] — we are looking for exactly such creators to be the face of our hair growth serum.

The serum is a roll-on with 3% Rosemary, 3% Redensyl, and 2% Anagain. Blocks DHT, reactivates dormant follicles, extends the growth phase. 8 to 12 weeks to visible results. Clean lab, 4.1 stars across 2,376 reviews on Amazon. Already the #8 best seller on TikTok in beauty, with creators bagging upwards of $700K on this product alone.

We are putting together a small, closed group of creators who have real authority in the hair space, with 40% commissions (well above category standard) and the chance to co-create a formulation with us over the next year. We are keeping the group small on purpose, and you are on the shortlist. We also have a limited number of retainer spots available, allocated on a first-come basis.

Send across your address and phone number and I will get a sample dispatched today.

Looking forward to hearing from you.

Mayank
CEO, Root Labs & Be Bodywise
---

RULES FOR [X] AND [Y]:

1. [X] = a specific reference. Examples: "your [topic] video", "your [routine name] breakdown", "your page", "your [product type] review". If transcript is rich, use a specific video reference. If transcript is empty/sparse, reference "your page" and lean on bio/niche.

2. [Y] = something concrete and specific you liked. Examples that worked in past sends:
   - "how cleanly you broke down Rosemary as the answer for thickness"
   - "the way you walked viewers through the three signs before pitching anything"
   - "that you do not just review the unit, you give it a vibe"
   - "your delivery, sharp, confident, zero filler"
   - "how openly you talk about your own hair concerns"
   - "that mechanism-led teaching is exactly the educator voice we are looking for"

3. CRITICAL TONE RULES (these override everything):
   - NO em dashes anywhere in YOUR custom additions. The template already contains two em dashes; do not add more.
   - NO contractions in YOUR custom additions ("I am" not "I'm", "you are" not "you're", "do not" not "don't").
   - The compliment must read as if Mayank actually saw the content. Avoid generic flattery.
   - Do NOT make the [X][Y] line a product pitch. Compliment the creator's voice/style/craft, not their product fit.

4. WHAT TO COMPLIMENT (in order of preference):
   - Specific moments from the transcript (a line, a framing, a structural choice)
   - Their delivery style (sharp, educational, mechanism-led, self-aware, etc.)
   - Their credentials if mentioned in bio (cosmetologist, esthetician, salon owner, etc.)
   - Their niche discipline (sticking to one lane in a noisy category)
   - Last resort: the bio voice itself ("just a girl who likes to do her hair")

5. IF VIDEO IS OFF-CATEGORY (e.g., skincare or body care when their bio says hair): do NOT reference the video. Use "your page" as [X] and compliment broader signals from bio/niche.

6. IF CREATOR APPEARS TO BE A FOUNDER/CEO of a competing hair growth brand (check bio for "CEO of", "founder of", and brand-name handles like @miraclegrowthwater): set status to "DO_NOT_SEND" and explain in reasoning. Do NOT draft the email.

OUTPUT FORMAT (return ONLY this JSON, no preamble):

{{
  "status": "ok" | "DO_NOT_SEND",
  "subject": "Be Bodywise x @{handle} — hair serum partnership",
  "body": "the full email body following the template exactly",
  "confidence": "high" | "medium" | "low",
  "reasoning": "one sentence on what X and Y were grounded in (e.g., 'X = her diagnostic framing video, Y = the three-sign structure she used before pitching')"
}}

Confidence guide:
  high   = strong specific transcript-grounded reference
  medium = bio-grounded or general delivery compliment
  low    = thin signal, bio is generic, no good handle on creator voice

CREATOR DATA:
  Handle: @{handle}
  First name: {first_name}
  Display name: {display_name}
  Niche: {niche}
  Bio: {bio}

VIDEO DATA:
  Video URL: {video_url}
  Transcript: {transcript}
  Visual summary: {visual_summary}
"""


def draft_email(creator: dict, video_url: str, transcript: str, visual_summary: str, anthropic: Anthropic) -> dict:
    # Pull a first name from display name
    display_name = creator.get("display_name", "") or ""
    first_name = display_name.strip().split()[0] if display_name.strip() else "there"
    # Strip emoji/special chars from first name
    first_name = re.sub(r"[^\w\s'-]", "", first_name).strip() or "there"

    prompt = EMAIL_PROMPT.format(
        handle=creator["handle"].lstrip("@"),
        first_name=first_name,
        display_name=display_name,
        niche=creator.get("niche", ""),
        bio=creator.get("bio", ""),
        video_url=video_url or "(no video available)",
        transcript=transcript or "(no transcript)",
        visual_summary=visual_summary or "(no visual summary)",
    )

    try:
        message = anthropic.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )
        text = message.content[0].text.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.MULTILINE).strip()
        return json.loads(text)
    except Exception as e:
        return {
            "status": "error",
            "subject": "",
            "body": "",
            "confidence": "low",
            "reasoning": f"Generation error: {e}",
        }


# =========================
# Main pipeline
# =========================

def run_pipeline(input_csv: str, output_path: str):
    anthropic = Anthropic(api_key=ANTHROPIC_API_KEY)

    # Read input
    with open(input_csv, "r", encoding="utf-8") as f:
        creators = list(csv.DictReader(f))
    print(f"[Input] {len(creators)} creators loaded")

    # Step 1: batch fetch all videos
    handles = [c["handle"] for c in creators]
    videos_by_handle = fetch_videos_for_handles(handles)

    # Step 2-4: per-creator processing
    results = []
    for i, creator in enumerate(creators, 1):
        handle_clean = creator["handle"].lstrip("@").lower()
        print(f"\n[{i}/{len(creators)}] @{handle_clean}")

        videos = videos_by_handle.get(handle_clean, [])
        if not videos:
            print("  no videos returned — falling back to bio-only")
            email = draft_email(creator, "", "", "", anthropic)
            results.append({**creator, "video_url": "", **email})
            continue

        best = select_best_video(videos)
        if not best:
            print("  no qualifying videos — falling back to bio-only")
            email = draft_email(creator, "", "", "", anthropic)
            results.append({**creator, "video_url": "", **email})
            continue

        video_url = best.get("webVideoUrl") or best.get("url") or ""
        print(f"  picked: {video_url}")

        ts = get_transcript(video_url, anthropic)
        transcript = ts.get("transcript", "") or ""
        visual_summary = ts.get("visual_summary", "") or ""
        if ts.get("error"):
            print(f"  transcript error: {ts['error'][:120]}")

        email = draft_email(creator, video_url, transcript, visual_summary, anthropic)
        results.append({**creator, "video_url": video_url, **email})
        print(f"  status={email.get('status')} confidence={email.get('confidence')}")

        time.sleep(1)  # gentle pacing

    # Write output to XLSX with two sheets: Outreach (mail-merge ready) + Review (full context)
    write_xlsx_output(results, output_path)

    print(f"\n[Output] {len(results)} rows written to {output_path}")
    print("Confidence breakdown:")
    from collections import Counter
    counts = Counter(r.get("confidence", "?") for r in results)
    for k, v in counts.items():
        print(f"  {k}: {v}")

    # Push the output XLSX back to the GitHub repo so it can be retrieved
    commit_output_to_github(output_path)


# =========================
# XLSX writer
# =========================

OUTREACH_HEADERS = ["Email Address", "Name", "Subject", "Body", "Image", "Status", "Sent At", "From"]
REVIEW_HEADERS = ["Handle", "Email Address", "Name", "Niche", "Bio", "Video URL", "Confidence", "Reasoning", "Status"]


def write_xlsx_output(results: list[dict], output_path: str):
    wb = Workbook()

    # Sheet 1: Outreach (mail-merge ready, exactly the 8 required columns)
    ws_out = wb.active
    ws_out.title = "Outreach"
    ws_out.append(OUTREACH_HEADERS)

    for r in results:
        gen_status = r.get("status", "")
        # Map generation status -> mail-merge Status column
        if gen_status == "ok":
            mail_status = "Pending"
        elif gen_status == "DO_NOT_SEND":
            mail_status = "DO_NOT_SEND"
        elif gen_status == "error":
            mail_status = "Error"
        else:
            mail_status = "Pending"

        ws_out.append([
            r.get("email", ""),
            r.get("display_name", ""),
            r.get("subject", ""),
            r.get("body", ""),
            "",  # Image — left blank, user fills if needed
            mail_status,
            "",  # Sent At — populated by mail merge tool
            "",  # From — set globally in mail merge tool
        ])

    style_sheet(ws_out, OUTREACH_HEADERS, body_col_idx=4, wide_cols={1: 32, 2: 22, 3: 50, 4: 80, 6: 14})

    # Sheet 2: Review (internal context for QA)
    ws_rev = wb.create_sheet("Review")
    ws_rev.append(REVIEW_HEADERS)

    for r in results:
        ws_rev.append([
            r.get("handle", ""),
            r.get("email", ""),
            r.get("display_name", ""),
            r.get("niche", ""),
            r.get("bio", ""),
            r.get("video_url", ""),
            r.get("confidence", ""),
            r.get("reasoning", ""),
            r.get("status", ""),
        ])

    style_sheet(ws_rev, REVIEW_HEADERS, wide_cols={1: 22, 2: 30, 3: 22, 4: 18, 5: 50, 6: 50, 7: 12, 8: 60})

    wb.save(output_path)


def style_sheet(ws, headers: list[str], body_col_idx: Optional[int] = None, wide_cols: Optional[dict] = None):
    # Header row formatting
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    header_align = Alignment(horizontal="left", vertical="center")

    for i, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Default column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Custom widths
    if wide_cols:
        for col_idx, width in wide_cols.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Body column: wrap text and consistent font for all data cells
    body_font = Font(name="Arial", size=11)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"


# =========================
# Push output back to GitHub repo (so Railway runs are retrievable)
# =========================

def commit_output_to_github(xlsx_path: str):
    """
    Pushes the XLSX to outputs/ folder in the GitHub repo via the GitHub API.
    Requires GITHUB_TOKEN (PAT with repo write scope) and GITHUB_REPO ("user/repo").
    """
    import base64
    import json as _json
    import urllib.request
    import urllib.error

    token = os.environ.get("GITHUB_TOKEN")
    repo = os.environ.get("GITHUB_REPO")

    if not all([token, repo]):
        print("[Git] Skipped — set GITHUB_TOKEN and GITHUB_REPO to enable")
        return

    with open(xlsx_path, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode()

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"creators_output_{timestamp}.xlsx"
    path_in_repo = f"outputs/{filename}"

    url = f"https://api.github.com/repos/{repo}/contents/{path_in_repo}"
    payload = {
        "message": f"Outreach run {timestamp}",
        "content": content_b64,
    }

    req = urllib.request.Request(
        url,
        data=_json.dumps(payload).encode(),
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json",
        },
        method="PUT",
    )

    try:
        with urllib.request.urlopen(req) as response:
            response.read()
            print(f"[Git] Pushed outputs/{filename} to {repo}")
    except urllib.error.HTTPError as e:
        print(f"[Git] Push failed: {e.code} {e.read().decode()[:200]}")
    except Exception as e:
        print(f"[Git] Push failed: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="creators_input.csv")
    parser.add_argument("--output", default="creators_output.xlsx")
    args = parser.parse_args()
    run_pipeline(args.input, args.output)
